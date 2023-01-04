from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.utils import get_column_letter
excel_name = input("ddl을 생성할 excel파일명(확장자는 빼주세요)\n")
excel_name.replace(' ', '')
print(excel_name)
string = ('C:/Users/minho/Desktop/MEARI_project/스마트검침 산출물/테이블정의서 버전관리/테이블기술서_v0.1/' + excel_name + '.xlsx')
wb = load_workbook(string)
#wb = load_workbook('C:/Users/minho/Desktop/MEARI_project/스마트검침 산출물/테이블정의서 버전관리/테이블기술서_v0.1/05.단말기계약납품관리_v0.1.xlsx')
ws = wb.active

get_column_name = ws['C10': 'C'+str(ws.max_row)]
get_Type = ws['F10': 'F'+str(ws.max_row)]
get_null = ws['G10': 'G'+str(ws.max_row)]
get_length_int = ws['H10': 'H' + str(ws.max_row)]
get_length_point = ws['I10': 'I' + str(ws.max_row)]
get_pk = ws['J10': 'J' + str(ws.max_row)]
get_default = ws['Q10': 'Q' + str(ws.max_row)]

columnName = []
Type = []
null = []
length_int = []
length_point = []
pk = []
default = []
pk_name = []


file_name = input("저장할 파일명:")
table_name = input("테이블명:")

f = open("C:/Users/minho/AppData/Roaming/DBeaverData/workspace6/General/Scripts/" + file_name + ".sql", 'w')

#컬럼명 리스트에 저장
for row in get_column_name:
    for cell in row:
        columnName.append(cell.value)
#데이터타입 리스트에 저장
for row in get_Type:
    for cell in row:
        Type.append(cell.value)
#null 여부 리스트에 저장
for row in get_null:
    for cell in row:
        if cell.value is None:
            null.append('')
        else:
            null.append('NOT NULL')
#정수 길이 리스트에 저장
for row in get_length_int:
    for cell in row:
        if cell.value is None:
            length_int.append('')
        else:
            length_int.append(cell.value)
#소수점 길이 리스트에 저장
for row in get_length_point:
    for cell in row:
        if cell.value is None:
            length_point.append('0')
        else:
            length_point.append(cell.value)
#PK 여부 리스트에 저장
for row in get_pk:
    for cell in row:
        # if cell.value is None:
        #     pk.append('')
        # else:
        pk.append(cell.value)
#default값 리스트에 저장
for row in get_default:
    for cell in row:
        if cell.value is None:
            default.append('')
        elif 'DEFAULT' in cell.value:
            default.append(cell.value.replace('= ', ''))
        else:
            default.append('')
# print(length_int)
# print(len(length_int))
# print(length_point)
# print(len(length_point))
# print(pk)
# print(len(pk))
# print(default)
# print(len(default))

f.write("CREATE TABLE " + table_name + " (\n")
for i in range(len(columnName)):
    f.write(columnName[i] + ' ' + Type[i])
    if Type[i] == 'DATE':
        f.write(' ')
    elif Type[i] == 'NUMBER':
        f.write('(' + str(length_int[i]) + ',' + str(length_point[i]) + ') ')
    else:
        f.write(('(' + str(length_int[i]) + ') '))
    f.write(default[i] + ' ' + null[i] + ',\n')
    if pk[i] is not None:
        pk_name.append(columnName[i])

f.write('\nCONSTRAINT ' + table_name + '_PK PRIMARY KEY(')
for i in range(len(pk_name)):
    if i == len(pk_name)-1:
        f.write(pk_name[i]+')\n')
    else:
        f.write(pk_name[i]+',')
f.write(");")
